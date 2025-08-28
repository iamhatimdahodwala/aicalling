from __future__ import annotations

from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException, UploadFile, File, Request, Body
from pydantic import BaseModel
from openpyxl import load_workbook

from vapi.types.create_customer_dto import CreateCustomerDto
from vapi.types.schedule_plan import SchedulePlan
from vapi.types.assistant_overrides import AssistantOverrides

from ..services.vapi_client import get_vapi_client_from_request
from ..config import settings
import httpx


router = APIRouter(prefix="/api/calls", tags=["calls"])


@router.get("")
def list_calls(limit: int = 100, status: Optional[str] = None, request: Request = None) -> List[Dict[str, Any]]:
	client = get_vapi_client_from_request(request)
	calls = client.calls.list(limit=limit)
	items = [c.dict() for c in calls]
	if status:
		status_lower = status.lower()
		items = [c for c in items if (c.get("status") or "").lower() == status_lower]
	return items


@router.get("/{call_id}")
def get_call(call_id: str, request: Request) -> Dict[str, Any]:
	client = get_vapi_client_from_request(request)
	try:
		call = client.calls.get(call_id)
		return call.dict()
	except Exception as e:
		raise HTTPException(status_code=404, detail=str(e))


@router.get("/{call_id}/artifacts")
def get_call_artifacts(call_id: str, request: Request) -> Dict[str, Any]:
	client = get_vapi_client_from_request(request)
	call = client.calls.get(call_id)
	artifact = (call.artifact or {}).dict() if hasattr(call, "artifact") and call.artifact is not None else {}
	return {
		"transcript": artifact.get("transcript"),
		"recordingUrl": artifact.get("recordingUrl"),
		"stereoRecordingUrl": artifact.get("stereoRecordingUrl"),
		"videoRecordingUrl": artifact.get("videoRecordingUrl"),
		"recording": artifact.get("recording"),
	}


@router.post("/{call_id}/terminate")
def terminate_call(call_id: str, request: Request) -> Dict[str, Any]:
	"""Terminate a call if possible by deleting it.

	Depending on provider/state, this may end the active call or remove the record.
	"""
	client = get_vapi_client_from_request(request)
	try:
		resp = client.calls.delete(call_id)
		return resp.dict()
	except Exception as e:
		raise HTTPException(status_code=400, detail=str(e))


@router.post("/{call_id}/escalate")
def escalate_call(call_id: str, destination: Optional[str] = None) -> Dict[str, Any]:
	"""Escalate by POST-ing to org webhook if configured.

	Set ESCALATE_WEBHOOK_URL in env to forward this request with callId and destination.
	"""
	payload = {"callId": call_id, "destination": destination}
	if settings.ESCALATE_WEBHOOK_URL:
		try:
			with httpx.Client(timeout=10) as client:
				client.post(settings.ESCALATE_WEBHOOK_URL, json=payload)
		except Exception as e:
			raise HTTPException(status_code=502, detail=f"Escalate webhook failed: {e}")
	return {"ok": True, **payload}


@router.post("/{call_id}/context")
def update_live_context(call_id: str, request: Request, body: Any = Body(...)) -> Dict[str, Any]:
	"""Update context during live call.

	Note: Real-time context updates are typically handled via the monitor control API/websocket.
	This endpoint stores your intent and can be wired to your server webhook or control channel.
	"""
	# Accept either a raw JSON string or an object with { text }
	if isinstance(body, str):
		text = body
	elif isinstance(body, dict):
		text = body.get("text") or body.get("message") or ""
	else:
		text = ""
	return {"ok": True, "callId": call_id, "text": text}


@router.get("/schedule/template")
def download_schedule_template() -> Dict[str, Any]:
	"""Return a simple template description for Excel columns.

	Frontend will use this to render or generate a file; alternatively you can serve a static .xlsx later.
	"""
	return {
		"columns": [
			{"key": "name", "example": "John Smith"},
			{"key": "number", "example": "+14155551234"},
			{"key": "earliest_at", "example": "2025-08-27T13:00:00Z"},
			{"key": "latest_at", "example": "2025-08-27T16:00:00Z"},
		],
	}


@router.post("/schedule/upload")
async def schedule_calls_from_excel(
	assistant_id: str,
	request: Request,
	file: UploadFile = File(...),
):
	"""Upload Excel with headers: name, number, earliest_at, latest_at (ISO8601).

	Creates batch outbound calls via client.calls.create(customers=[...], schedule_plan=...).
	"""
	client = get_vapi_client_from_request(request)
	content = await file.read()
	wb = load_workbook(filename=(file.filename or "uploaded.xlsx"), data_only=True)
	# openpyxl requires a file path or file-like object; since FastAPI gives bytes, reopen via BytesIO
	# To avoid extra deps, write to temp file
	import tempfile
	with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
		tmp.write(content)
		tmp.flush()
		wb = load_workbook(filename=tmp.name, data_only=True)
	ws = wb.active
	rows = list(ws.iter_rows(values_only=True))
	if not rows:
		raise HTTPException(status_code=400, detail="Empty Excel file")
	headers = [str(h).strip().lower() for h in rows[0]]
	try:
		name_idx = headers.index("name")
		number_idx = headers.index("number")
		earliest_idx = headers.index("earliest_at")
		latest_idx = headers.index("latest_at") if "latest_at" in headers else None
	except ValueError:
		raise HTTPException(status_code=400, detail="Headers must include name, number, earliest_at, optional latest_at")

	customers: List[CreateCustomerDto] = []
	for row in rows[1:]:
		if row is None:
			continue
		name = str(row[name_idx]) if row[name_idx] is not None else None
		number = str(row[number_idx]) if row[number_idx] is not None else None
		if not number:
			continue
		customers.append(CreateCustomerDto(name=name, number=number))

	if not customers:
		raise HTTPException(status_code=400, detail="No valid rows found")

	# For batch scheduling, we set a broad schedule plan from the first row.
	first = rows[1]
	earliest_at_raw = first[earliest_idx]
	latest_at_raw = first[latest_idx] if latest_idx is not None else None

	def parse_dt(val: Any) -> datetime:
		if isinstance(val, datetime):
			return val
		return datetime.fromisoformat(str(val))

	schedule = SchedulePlan(
		earliest_at=parse_dt(earliest_at_raw),
		latest_at=parse_dt(latest_at_raw) if latest_at_raw else None,
	)

	resp = client.calls.create(
		assistant_id=assistant_id,
		customers=customers,
		schedule_plan=schedule,
	)
	return resp.dict()


class ScheduleSingleBody(BaseModel):
	assistant_id: str
	name: str | None = None
	number: str
	earliest_at: datetime
	latest_at: datetime | None = None
	context: str | None = None


class StartWebCallBody(BaseModel):
	assistant_id: str
	customer_name: str | None = None
	# Optionally future: metadata, variables


@router.post("/schedule/single")
def schedule_single(body: ScheduleSingleBody, request: Request) -> Dict[str, Any]:
	"""Schedule a single outbound call for a customer.

	Body: { assistant_id, name?, number, earliest_at, latest_at? }
	"""
	client = get_vapi_client_from_request(request)
	schedule = SchedulePlan(
		earliest_at=body.earliest_at,
		latest_at=body.latest_at,
	)
	customer = CreateCustomerDto(name=body.name, number=body.number)
	resp = client.calls.create(
		assistant_id=body.assistant_id,
		customers=[customer],
		schedule_plan=schedule,
		assistant_overrides=AssistantOverrides(variable_values={"context": body.context} if body.context else None),
	)
	return resp.dict()


@router.post("/web/start")
def start_web_call(body: StartWebCallBody, request: Request) -> Dict[str, Any]:
	"""Start a web call for coaching/training.

	Creates a call with a special 'web' target if supported by SDK.
	"""
	client = get_vapi_client_from_request(request)
	try:
		create_kwargs = {
			"assistant_id": body.assistant_id,
			"customer": CreateCustomerDto(name=body.customer_name or "Web User", number=None),
		}
		# Try common flag for web session
		for alt in (
			{"web": {"enabled": True}},
			{"channel": "web"},
		):
			try:
				resp = client.calls.create(**create_kwargs, **alt)
				return resp.dict()
			except Exception:
				continue
		raise HTTPException(status_code=501, detail="Web call creation not supported by SDK")
	except Exception as e:
		raise HTTPException(status_code=400, detail=str(e))


@router.post("/web/end/{call_id}")
def end_web_call(call_id: str, request: Request) -> Dict[str, Any]:
	client = get_vapi_client_from_request(request)
	try:
		resp = client.calls.delete(call_id)
		return resp.dict()
	except Exception as e:
		raise HTTPException(status_code=400, detail=str(e))

